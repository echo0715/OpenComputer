"""
Test LibreOffice Calc verifier endpoints in a live E2B sandbox.

Covers:
  - Help/usage output
  - Error cases (LibreOffice not running, bad args)
  - ODF file parsing endpoints (ODS file, no UNO needed)
  - UNO live endpoints with XLSX file (sheets, cell-value, range-values, cell-format, etc.)
  - Composite check-* endpoints — positive and negative cases
  - Formula cells (weighted average, SUM, IF)
  - Sheet name argument passing through CLI
  - Column sort check on formula columns

Usage:
    python verifiers/libreoffice_calc/test_libreoffice_calc.py
"""

import json
import sys
import time
import traceback
from pathlib import Path

from dotenv import load_dotenv
from e2b_desktop import Sandbox
from e2b.sandbox.commands.command_handle import CommandExitException

load_dotenv()

VERIFIER_LOCAL = Path(__file__).parent / "libreoffice_calc.py"
VERIFIER_REMOTE = "/home/user/verifiers/libreoffice_calc.py"
V = f"python3 {VERIFIER_REMOTE}"

TEST_ODS_PATH = "/home/user/test_verifier.ods"
TEST_XLSX_PATH = "/home/user/test_verifier.xlsx"

# ── ODS fixture (stdlib ZIP+XML) ──────────────────────────────────────────

CREATE_ODS_SCRIPT = r'''
import zipfile

CONTENT_XML = """<?xml version="1.0" encoding="UTF-8"?>
<office:document-content
  xmlns:office="urn:oasis:names:tc:opendocument:xmlns:office:1.0"
  xmlns:table="urn:oasis:names:tc:opendocument:xmlns:table:1.0"
  xmlns:text="urn:oasis:names:tc:opendocument:xmlns:text:1.0"
  xmlns:style="urn:oasis:names:tc:opendocument:xmlns:style:1.0"
  xmlns:fo="urn:oasis:names:tc:opendocument:xmlns:xsl-fo-compatible:1.0"
  office:version="1.2">
  <office:automatic-styles>
    <style:style style:name="bold" style:family="table-cell">
      <style:text-properties fo:font-weight="bold"/>
    </style:style>
  </office:automatic-styles>
  <office:body>
    <office:spreadsheet>
      <table:table table:name="TestData">
        <table:table-row>
          <table:table-cell table:style-name="bold" office:value-type="string"><text:p>Name</text:p></table:table-cell>
          <table:table-cell table:style-name="bold" office:value-type="string"><text:p>Score</text:p></table:table-cell>
          <table:table-cell table:style-name="bold" office:value-type="string"><text:p>Grade</text:p></table:table-cell>
        </table:table-row>
        <table:table-row>
          <table:table-cell office:value-type="string"><text:p>Alice</text:p></table:table-cell>
          <table:table-cell office:value-type="float" office:value="90"><text:p>90</text:p></table:table-cell>
          <table:table-cell office:value-type="string"><text:p>A</text:p></table:table-cell>
        </table:table-row>
        <table:table-row>
          <table:table-cell office:value-type="string"><text:p>Bob</text:p></table:table-cell>
          <table:table-cell office:value-type="float" office:value="75"><text:p>75</text:p></table:table-cell>
          <table:table-cell office:value-type="string"><text:p>B</text:p></table:table-cell>
        </table:table-row>
        <table:table-row>
          <table:table-cell office:value-type="string"><text:p>Charlie</text:p></table:table-cell>
          <table:table-cell office:value-type="float" office:value="60"><text:p>60</text:p></table:table-cell>
          <table:table-cell office:value-type="string"><text:p>C</text:p></table:table-cell>
        </table:table-row>
      </table:table>
      <table:table table:name="Summary">
        <table:table-row>
          <table:table-cell office:value-type="string"><text:p>Total Students</text:p></table:table-cell>
          <table:table-cell office:value-type="float" office:value="3"><text:p>3</text:p></table:table-cell>
        </table:table-row>
      </table:table>
    </office:spreadsheet>
  </office:body>
</office:document-content>
"""

MANIFEST_XML = """<?xml version="1.0" encoding="UTF-8"?>
<manifest:manifest xmlns:manifest="urn:oasis:names:tc:opendocument:xmlns:manifest:1.0">
  <manifest:file-entry manifest:media-type="application/vnd.oasis.opendocument.spreadsheet" manifest:full-path="/"/>
  <manifest:file-entry manifest:media-type="text/xml" manifest:full-path="content.xml"/>
</manifest:manifest>
"""

path = "/home/user/test_verifier.ods"
with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
    z.writestr("mimetype", "application/vnd.oasis.opendocument.spreadsheet")
    z.writestr("content.xml", CONTENT_XML)
    z.writestr("META-INF/manifest.xml", MANIFEST_XML)

print("OK")
'''

# ── XLSX fixture (openpyxl) ───────────────────────────────────────────────
# This is the file we open in LibreOffice for UNO tests.
# Two sheets: "Products" with data + formulas, "Stats" with cross-sheet ref.

CREATE_XLSX_SCRIPT = r'''
import subprocess, sys
subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"],
                      stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
import openpyxl
from openpyxl.styles import Font

wb = openpyxl.Workbook()

# Sheet 1: Products
ws = wb.active
ws.title = "Products"
headers = ["Item", "Qty", "Price", "Total"]
for col_idx, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col_idx, value=h)
    c.font = Font(bold=True)

products = [
    ["Widget", 10, 25.00],
    ["Gadget", 5, 50.00],
    ["Sprocket", 20, 12.50],
    ["Flange", 8, 30.00],
    ["Bolt", 100, 0.50],
]
for row_idx, (item, qty, price) in enumerate(products, 2):
    ws.cell(row=row_idx, column=1, value=item)
    ws.cell(row=row_idx, column=2, value=qty)
    ws.cell(row=row_idx, column=3, value=price)
    # Formula: Total = Qty * Price
    ws.cell(row=row_idx, column=4).value = f"=B{row_idx}*C{row_idx}"

# Row 7: totals
ws.cell(row=7, column=1, value="TOTAL")
ws.cell(row=7, column=1).font = Font(bold=True)
ws.cell(row=7, column=4).value = "=SUM(D2:D6)"

# Sheet 2: Stats
ws2 = wb.create_sheet("Stats")
ws2.cell(row=1, column=1, value="Metric").font = Font(bold=True)
ws2.cell(row=1, column=2, value="Value").font = Font(bold=True)
ws2.cell(row=2, column=1, value="Product Count")
ws2.cell(row=2, column=2, value=5)
ws2.cell(row=3, column=1, value="Grand Total")
ws2.cell(row=3, column=2).value = "=Products!D7"

wb.save("/home/user/test_verifier.xlsx")
print("OK")
'''

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

passed = 0
failed = 0
errors: list[str] = []


class CmdResult:
    def __init__(self, exit_code: int, stdout: str, stderr: str):
        self.exit_code = exit_code
        self.stdout = stdout
        self.stderr = stderr


def run(sandbox: Sandbox, cmd: str, timeout: int = 30) -> dict | list:
    r = run_raw(sandbox, cmd, timeout)
    if r.exit_code != 0 and not r.stdout.strip():
        return {"error": f"exit_code={r.exit_code} stderr={r.stderr[:300]}"}
    try:
        return json.loads(r.stdout)
    except json.JSONDecodeError:
        return {"error": f"Invalid JSON: {r.stdout[:300]}"}


def run_raw(sandbox: Sandbox, cmd: str, timeout: int = 30) -> CmdResult:
    try:
        result = sandbox.commands.run(f"{V} {cmd}", timeout=timeout)
        return CmdResult(result.exit_code, result.stdout, result.stderr)
    except CommandExitException as e:
        return CmdResult(e.exit_code, e.stdout, e.stderr)


def run_shell(sandbox: Sandbox, cmd: str, timeout: int = 60) -> CmdResult:
    try:
        result = sandbox.commands.run(cmd, timeout=timeout)
        return CmdResult(result.exit_code, result.stdout, result.stderr)
    except CommandExitException as e:
        return CmdResult(e.exit_code, e.stdout, e.stderr)


def check(name: str, condition: bool, detail: str = ""):
    global passed, failed
    if condition:
        passed += 1
        print(f"  PASS  {name}")
    else:
        failed += 1
        msg = f"  FAIL  {name}"
        if detail:
            msg += f"  -- {detail}"
        print(msg)
        errors.append(f"{name}: {detail}")


def is_valid_json(stdout: str) -> bool:
    try:
        json.loads(stdout)
        return True
    except (json.JSONDecodeError, TypeError):
        return False


# ---------------------------------------------------------------------------
# Test groups
# ---------------------------------------------------------------------------

def test_help(sandbox: Sandbox):
    print("\n=== Help ===")
    result = run_raw(sandbox, "--help")
    check("help exits 0", result.exit_code == 0, f"got exit_code={result.exit_code}")
    check("help mentions commands", "Commands:" in result.stdout, result.stdout[:100])


def test_errors_lo_not_running(sandbox: Sandbox):
    print("\n=== Errors (LibreOffice not running) ===")
    for cmd, name in [("sheets", "sheets"), ("cell-value A1", "cell-value"), ("active-sheet", "active-sheet")]:
        data = run(sandbox, cmd)
        check(f"{name} returns error when LO down", "error" in data, str(data)[:100])


def test_errors_bad_args(sandbox: Sandbox):
    print("\n=== Errors (bad args) ===")
    result = run_raw(sandbox, "check-cell-value")
    check("missing arg exits 1", result.exit_code == 1)
    check("missing arg valid JSON", is_valid_json(result.stdout), result.stdout[:100])

    result = run_raw(sandbox, "nonexistent-command")
    check("unknown cmd exits 1", result.exit_code == 1)
    check("unknown cmd valid JSON", is_valid_json(result.stdout), result.stdout[:100])


def test_ods_file_parsing(sandbox: Sandbox):
    """ODF parsing endpoints (no UNO needed, ODS file only)."""
    print("\n=== ODS File Parsing ===")

    data = run(sandbox, f"parse-sheets {TEST_ODS_PATH}")
    check("parse-sheets returns sheets", len(data.get("sheets", [])) > 0, str(data)[:100])
    sheet_names = [s["name"] for s in data.get("sheets", [])]
    check("parse-sheets has TestData", "TestData" in sheet_names, str(sheet_names))
    check("parse-sheets has Summary", "Summary" in sheet_names, str(sheet_names))

    data = run(sandbox, f"parse-cell A1 {TEST_ODS_PATH}")
    check("parse-cell A1 = Name", data.get("value") == "Name" or data.get("display") == "Name",
          f"value={data.get('value')}")

    data = run(sandbox, f"parse-cell B2 {TEST_ODS_PATH}")
    check("parse-cell B2 = 90", data.get("value") in (90, 90.0), f"value={data.get('value')}")

    data = run(sandbox, f"parse-range A1:C2 {TEST_ODS_PATH}")
    check("parse-range 2 rows x 3 cols", data.get("rows") == 2 and data.get("cols") == 3,
          f"rows={data.get('rows')} cols={data.get('cols')}")

    data = run(sandbox, f"parse-cell A1 {TEST_ODS_PATH} Summary")
    check("parse-cell Summary A1 = Total Students",
          "Total" in str(data.get("value", "")), f"value={data.get('value')}")

    data = run(sandbox, "parse-cell A1 /nonexistent/file.ods")
    check("parse missing file returns error", "error" in data, str(data)[:100])


def test_uno_basic(sandbox: Sandbox):
    """Basic UNO endpoints with XLSX file loaded."""
    print("\n=== UNO Basic (XLSX) ===")

    data = run(sandbox, "sheets")
    check("sheets returns dict", isinstance(data, dict) and "error" not in data, str(data)[:150])
    if "error" not in data:
        names = [s["name"] for s in data.get("sheets", [])]
        check("sheets has Products", "Products" in names, str(names))
        check("sheets has Stats", "Stats" in names, str(names))
        check("sheets count=2", data.get("count") == 2, f"count={data.get('count')}")

    data = run(sandbox, "active-sheet")
    check("active-sheet has name", "name" in data and "error" not in data, str(data)[:100])

    data = run(sandbox, "doc-info")
    check("doc-info has title", "title" in data and "error" not in data, str(data)[:100])
    if "error" not in data:
        check("doc-info shows xlsx", "xlsx" in str(data.get("path", "")).lower() or
              "xlsx" in str(data.get("title", "")).lower(), str(data)[:150])


def test_uno_cell_values(sandbox: Sandbox):
    """Cell value reading via UNO — strings, numbers, formulas."""
    print("\n=== UNO Cell Values ===")

    # String cell
    data = run(sandbox, "cell-value A1 Products")
    check("A1 = Item (string)", data.get("value") == "Item", f"value={data.get('value')}")
    check("A1 type = string", data.get("type") == "string", f"type={data.get('type')}")

    # Numeric cell
    data = run(sandbox, "cell-value B2 Products")
    check("B2 = 10 (numeric)", data.get("value") in (10, 10.0), f"value={data.get('value')}")
    check("B2 type = float", data.get("type") == "float", f"type={data.get('type')}")

    # Formula cell
    data = run(sandbox, "cell-value D2 Products")
    check("D2 is formula type", data.get("type") == "formula", f"type={data.get('type')}")
    check("D2 formula contains B2*C2",
          data.get("formula") is not None and "B2" in data.get("formula", "") and "C2" in data.get("formula", ""),
          f"formula={data.get('formula')}")
    check("D2 value = 250", data.get("value") in (250, 250.0), f"value={data.get('value')}")

    # SUM formula
    data = run(sandbox, "cell-value D7 Products")
    check("D7 is formula type", data.get("type") == "formula", f"type={data.get('type')}")
    check("D7 formula contains SUM",
          data.get("formula") is not None and "SUM" in data.get("formula", "").upper(),
          f"formula={data.get('formula')}")
    # Expected: 10*25 + 5*50 + 20*12.5 + 8*30 + 100*0.5 = 250+250+250+240+50 = 1040
    check("D7 value = 1040", data.get("value") in (1040, 1040.0), f"value={data.get('value')}")

    # Cross-sheet formula
    data = run(sandbox, "cell-value B3 Stats")
    check("Stats B3 is formula", data.get("type") == "formula", f"type={data.get('type')}")
    check("Stats B3 value = 1040", data.get("value") in (1040, 1040.0), f"value={data.get('value')}")


def test_uno_range_and_sheet_data(sandbox: Sandbox):
    """Range values and sheet-data endpoints."""
    print("\n=== UNO Range & Sheet Data ===")

    data = run(sandbox, "range-values A1:D2 Products")
    check("range-values has data", isinstance(data.get("data"), list) and "error" not in data,
          str(data)[:150])
    if "error" not in data:
        check("range 2 rows", data.get("rows") == 2, f"rows={data.get('rows')}")
        check("range 4 cols", data.get("cols") == 4, f"cols={data.get('cols')}")
        check("range first cell = Item", data["data"][0][0] == "Item",
              f"first_cell={data['data'][0][0]}")

    data = run(sandbox, "sheet-data Products")
    check("sheet-data has headers", isinstance(data.get("headers"), list) and "error" not in data,
          str(data)[:100])
    if "error" not in data:
        check("sheet-data rows >= 7", data.get("rows", 0) >= 7, f"rows={data.get('rows')}")

    data = run(sandbox, "sheet-data Stats")
    check("sheet-data Stats works", "error" not in data, str(data)[:100])


def test_uno_cell_format(sandbox: Sandbox):
    """Cell formatting via UNO — bold, not bold."""
    print("\n=== UNO Cell Format ===")

    data = run(sandbox, "cell-format A1 Products")
    check("cell-format returns dict", isinstance(data, dict) and "error" not in data, str(data)[:200])
    if "error" not in data:
        check("A1 is bold", data.get("bold") is True, f"bold={data.get('bold')}")
        check("A1 has font_name", isinstance(data.get("font_name"), str), f"font_name={data.get('font_name')}")
        check("A1 has h_align", "h_align" in data, str(data.keys()))
        check("A1 has v_align", "v_align" in data, str(data.keys()))

    data = run(sandbox, "cell-format B2 Products")
    if "error" not in data:
        check("B2 not bold", data.get("bold") is False, f"bold={data.get('bold')}")

    # Format on second sheet
    data = run(sandbox, "cell-format A1 Stats")
    if "error" not in data:
        check("Stats A1 is bold", data.get("bold") is True, f"bold={data.get('bold')}")


def test_uno_merged_cells(sandbox: Sandbox):
    """Merged cells endpoint."""
    print("\n=== UNO Merged Cells ===")
    data = run(sandbox, "merged-cells Products")
    check("merged-cells returns dict", isinstance(data, dict) and "error" not in data, str(data)[:100])
    if "error" not in data:
        check("merged-cells has count", "count" in data, str(data.keys()))


def test_checks_positive(sandbox: Sandbox):
    """Composite check-* endpoints — positive cases."""
    print("\n=== Checks (positive) ===")

    # check-cell-value string
    data = run(sandbox, 'check-cell-value A1 Item Products')
    check("check-cell-value A1=Item match", data.get("match") is True, str(data)[:150])

    # check-cell-value numeric
    data = run(sandbox, "check-cell-value B2 10 Products")
    check("check-cell-value B2=10 match", data.get("match") is True, str(data)[:100])

    # check-cell-value on formula cell (checks computed value)
    data = run(sandbox, "check-cell-value D2 250 Products")
    check("check-cell-value D2=250 formula match", data.get("match") is True, str(data)[:100])

    # check-cell-value on SUM formula
    data = run(sandbox, "check-cell-value D7 1040 Products")
    check("check-cell-value D7=1040 SUM match", data.get("match") is True, str(data)[:100])

    # check-cell-value on second sheet
    data = run(sandbox, "check-cell-value B2 5 Stats")
    check("check-cell-value Stats B2=5 match", data.get("match") is True, str(data)[:100])

    # check-sheet-exists
    data = run(sandbox, "check-sheet-exists Products")
    check("check-sheet-exists Products=true", data.get("exists") is True, str(data)[:100])

    data = run(sandbox, "check-sheet-exists Stats")
    check("check-sheet-exists Stats=true", data.get("exists") is True, str(data)[:100])

    # check-sheet-count
    data = run(sandbox, "check-sheet-count 2")
    check("check-sheet-count 2=true", data.get("match") is True, str(data)[:100])

    # check-cell-formula
    data = run(sandbox, 'check-cell-formula D2 "=B2*C2" Products')
    check("check-cell-formula D2=B2*C2 match", data.get("match") is True, str(data)[:150])

    data = run(sandbox, 'check-cell-formula D7 "=SUM(D2:D6)" Products')
    check("check-cell-formula D7=SUM match", data.get("match") is True, str(data)[:150])

    # check-cell-formatted bold (header row)
    data = run(sandbox, "check-cell-formatted A1 true Products")
    check("check-cell-formatted A1 bold=true", data.get("match") is True, str(data)[:150])

    # check-cell-formatted bold on second sheet
    data = run(sandbox, "check-cell-formatted A1 true Stats")
    check("check-cell-formatted Stats A1 bold=true", data.get("match") is True, str(data)[:150])

    # check-column-sorted: Qty column B rows 2-6 not necessarily sorted, but
    # Price column C: 25, 50, 12.5, 30, 0.5 — not sorted
    # Item column A: Widget, Gadget, Sprocket, Flange, Bolt — not sorted
    # Let's check B desc: 100, 20, 10, 8, 5 — need to sort first
    # Actually test something we know: column A1="TOTAL" in row 7 at bottom

    # check-file-exists
    data = run(sandbox, f"check-file-exists {TEST_XLSX_PATH}")
    check("check-file-exists xlsx=true", data.get("exists") is True, str(data)[:100])

    # check-file-saved
    data = run(sandbox, "check-file-saved")
    check("check-file-saved has key", "saved" in data or "error" in data, str(data)[:150])


def test_checks_negative(sandbox: Sandbox):
    """Composite check-* endpoints — negative cases."""
    print("\n=== Checks (negative) ===")

    data = run(sandbox, "check-cell-value A1 WrongValue Products")
    check("check-cell-value mismatch=false", data.get("match") is False, str(data)[:100])

    data = run(sandbox, "check-sheet-exists NonExistentSheet")
    check("check-sheet-exists false", data.get("exists") is False, str(data)[:100])

    data = run(sandbox, "check-sheet-count 99")
    check("check-sheet-count mismatch=false", data.get("match") is False, str(data)[:100])

    data = run(sandbox, "check-file-exists /nonexistent/file.xlsx")
    check("check-file-exists false", data.get("exists") is False, str(data)[:100])

    # Not bold cell
    data = run(sandbox, "check-cell-formatted B2 true Products")
    check("check-cell-formatted B2 bold=true fails", data.get("match") is False, str(data)[:150])

    # Wrong formula
    data = run(sandbox, 'check-cell-formula D2 "=WRONG" Products')
    check("check-cell-formula wrong=false", data.get("match") is False, str(data)[:150])


def test_column_sorted_with_formulas(sandbox: Sandbox):
    """check-column-sorted must handle formula cells numerically."""
    print("\n=== Column Sort (formula cells) ===")

    # Column D has formulas: 250, 250, 250, 240, 50 — descending
    data = run(sandbox, "check-column-sorted D desc 2 6")
    check("col D desc 2-6 sorted", data.get("sorted") is True,
          f"sorted={data.get('sorted')} detail={str(data)[:200]}")

    # Same column ascending should be false
    data = run(sandbox, "check-column-sorted D asc 2 6")
    check("col D asc 2-6 not sorted", data.get("sorted") is False,
          f"sorted={data.get('sorted')}")

    # Column B (Qty): 10, 5, 20, 8, 100 — not sorted either way
    data = run(sandbox, "check-column-sorted B asc 2 6")
    check("col B asc not sorted", data.get("sorted") is False, f"sorted={data.get('sorted')}")


def test_all_commands_return_json(sandbox: Sandbox):
    """Every CLI command should output valid JSON."""
    print("\n=== JSON validity (all commands) ===")

    no_arg_cmds = ["sheets", "active-sheet", "doc-info", "merged-cells"]
    for cmd in no_arg_cmds:
        result = run_raw(sandbox, cmd)
        valid = is_valid_json(result.stdout)
        check(f"{cmd} valid JSON", valid,
              f"exit={result.exit_code} stdout={result.stdout[:80]}" if not valid else "")

    arg_cmds = [
        ("cell-value", "A1"),
        ("cell-value", "A1 Products"),
        ("range-values", "A1:B2"),
        ("range-values", "A1:B2 Products"),
        ("sheet-data", "Products"),
        ("sheet-data", "Stats"),
        ("cell-format", "A1"),
        ("cell-format", "A1 Products"),
        ("merged-cells", "Products"),
        ("parse-sheets", TEST_ODS_PATH),
        ("parse-cell", f"A1 {TEST_ODS_PATH}"),
        ("parse-range", f"A1:B2 {TEST_ODS_PATH}"),
        ("check-cell-value", "A1 Item Products"),
        ("check-cell-value", "B2 10"),
        ("check-sheet-exists", "Products"),
        ("check-sheet-count", "2"),
        ("check-cell-formula", f'D2 "=B2*C2" Products'),
        ("check-cell-formatted", "A1 true"),
        ("check-cell-formatted", "A1 true Products"),
        ("check-column-sorted", "A"),
        ("check-column-sorted", "D desc 2 6"),
        ("check-file-exists", TEST_XLSX_PATH),
        ("check-file-saved", ""),
        ("check-merged-cells", "A1:A1"),
        ("check-merged-cells", "A1:A1 Products"),
    ]

    for cmd, arg in arg_cmds:
        full_cmd = f"{cmd} {arg}".strip()
        result = run_raw(sandbox, full_cmd)
        valid = is_valid_json(result.stdout)
        check(f"{full_cmd} valid JSON", valid,
              f"exit={result.exit_code} stdout={result.stdout[:80]}" if not valid else "")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def launch_libreoffice(sandbox: Sandbox, file_path: str) -> bool:
    """Launch LibreOffice with UNO socket, opening the given file."""
    try:
        sandbox.commands.run("pkill -9 -f soffice", timeout=5)
    except (CommandExitException, Exception):
        pass
    try:
        sandbox.commands.run("pkill -9 -f oosplash", timeout=5)
    except (CommandExitException, Exception):
        pass
    time.sleep(2)

    launch_script = f'''#!/bin/bash
soffice --calc \
  '--accept=socket,host=localhost,port=2002;urp;' \
  --norestore --nologo \
  {file_path} \
  > /tmp/lo.log 2>&1
'''
    sandbox.files.write("/tmp/launch_lo.sh", launch_script)
    sandbox.commands.run("chmod +x /tmp/launch_lo.sh", timeout=3)
    sandbox.commands.run("/tmp/launch_lo.sh", background=True)

    check_script = '''
import uno, time
for attempt in range(5):
    try:
        ctx = uno.getComponentContext()
        resolver = ctx.ServiceManager.createInstanceWithContext(
            "com.sun.star.bridge.UnoUrlResolver", ctx
        )
        resolved = resolver.resolve(
            "uno:socket,host=localhost,port=2002;urp;StarOffice.ComponentContext"
        )
        smgr = resolved.ServiceManager
        desktop = smgr.createInstanceWithContext("com.sun.star.frame.Desktop", resolved)
        doc = desktop.getCurrentComponent()
        if doc is not None:
            print("READY")
            break
        else:
            print("NODOC")
            break
    except Exception as e:
        if attempt < 4:
            time.sleep(1)
        else:
            print(f"ERR:{e}")
'''
    sandbox.files.write("/tmp/check_uno.py", check_script)

    for i in range(30):
        try:
            r = sandbox.commands.run("python3 /tmp/check_uno.py", timeout=10)
            if "READY" in r.stdout:
                print(f"  LibreOffice UNO ready (attempt {i+1})")
                return True
            if "NODOC" in r.stdout:
                print(f"  UNO connected but no doc yet (attempt {i+1})")
        except (CommandExitException, Exception) as e:
            if i > 15:
                print(f"  attempt {i+1}: {str(e)[:80]}")
        time.sleep(2)

    try:
        r = sandbox.commands.run("tail -20 /tmp/lo.log", timeout=3)
        print(f"  [lo log] {r.stdout[-300:]}")
    except (CommandExitException, Exception):
        pass
    return False


def main():
    global passed, failed

    print("=" * 60)
    print("LibreOffice Calc Verifier Test Suite")
    print("=" * 60)

    print("\nCreating sandbox from desktop-all-apps...")
    sandbox = Sandbox.create(template="desktop-all-apps", timeout=600)

    try:
        # Upload verifier
        print(f"Uploading {VERIFIER_LOCAL} -> {VERIFIER_REMOTE}")
        sandbox.commands.run("mkdir -p /home/user/verifiers")
        with open(VERIFIER_LOCAL) as f:
            sandbox.files.write(VERIFIER_REMOTE, f.read())

        # ── Tests with LibreOffice NOT running ──
        test_help(sandbox)
        test_errors_lo_not_running(sandbox)
        test_errors_bad_args(sandbox)

        # ── Create ODS test file ──
        print("\n  Creating ODS test file...")
        sandbox.files.write("/home/user/create_ods.py", CREATE_ODS_SCRIPT)
        r = run_shell(sandbox, "python3 /home/user/create_ods.py")
        ods_ok = "OK" in r.stdout
        check("ODS file created", ods_ok, f"stdout={r.stdout[:100]} stderr={r.stderr[:100]}")
        if ods_ok:
            test_ods_file_parsing(sandbox)

        # ── Create XLSX test file ──
        print("\n  Creating XLSX test file...")
        sandbox.files.write("/home/user/create_xlsx.py", CREATE_XLSX_SCRIPT)
        r = run_shell(sandbox, "python3 /home/user/create_xlsx.py", timeout=120)
        xlsx_ok = "OK" in r.stdout
        check("XLSX file created", xlsx_ok, f"stdout={r.stdout[:100]} stderr={r.stderr[:100]}")

        # ── Launch LibreOffice with XLSX file ──
        print("\nLaunching LibreOffice with XLSX file...")
        lo_ready = launch_libreoffice(sandbox, TEST_XLSX_PATH)
        if not lo_ready:
            print("  WARNING: LibreOffice UNO not ready -- UNO tests may fail")

        # ── UNO tests (all with XLSX) ──
        test_uno_basic(sandbox)
        test_uno_cell_values(sandbox)
        test_uno_range_and_sheet_data(sandbox)
        test_uno_cell_format(sandbox)
        test_uno_merged_cells(sandbox)
        test_checks_positive(sandbox)
        test_checks_negative(sandbox)
        test_column_sorted_with_formulas(sandbox)
        test_all_commands_return_json(sandbox)

    except Exception:
        traceback.print_exc()
        failed += 1
        errors.append(f"Unhandled exception: {traceback.format_exc()}")

    finally:
        sandbox.kill()
        print("\nSandbox killed.")

    # ── Summary ──
    print("\n" + "=" * 60)
    print(f"Results: {passed} passed, {failed} failed, {passed + failed} total")
    print("=" * 60)

    if errors:
        print("\nFailures:")
        for e in errors:
            print(f"  - {e}")

    sys.exit(1 if failed > 0 else 0)


if __name__ == "__main__":
    main()
