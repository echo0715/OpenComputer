"""Generate env .xlsx files for libreoffice_calc v2 tasks.

Each function builds one xlsx, saves it into task_generator/tasks/<id>/env/,
writes env_manifest.json, and then re-opens the file with openpyxl to assert
content. Expected values for verification are printed to stdout for use in
task.json.
"""

import json
import os
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font

ROOT = Path(__file__).parent

BOLD = Font(bold=True)


def ensure(task_id: str) -> Path:
    p = ROOT / task_id / "env"
    p.mkdir(parents=True, exist_ok=True)
    return p


def save_manifest(task_id: str, files):
    manifest = {"task_id": task_id, "files": files}
    (ROOT / task_id / "env_manifest.json").write_text(json.dumps(manifest, indent=2))


def reopen_check(path: Path):
    wb = openpyxl.load_workbook(path, data_only=False)
    return wb


# ----- 1. calc_loan_amortization_schedule -----
def gen_loan():
    tid = "calc_loan_amortization_schedule"
    env = ensure(tid)
    wb = Workbook()
    ws = wb.active
    ws.title = "Loan"
    ws["A1"] = "Principal"; ws["B1"] = 24000
    ws["A2"] = "Annual Rate"; ws["B2"] = 0.06
    ws["A3"] = "Term Months"; ws["B3"] = 12
    ws["A1"].font = BOLD; ws["A2"].font = BOLD; ws["A3"].font = BOLD
    ws["A5"] = "Amortization Schedule"; ws["A5"].font = BOLD
    headers = ["Month", "Payment", "Interest", "Principal", "Balance"]
    for i, h in enumerate(headers):
        c = ws.cell(row=6, column=1 + i, value=h); c.font = BOLD
    path = env / "loan.xlsx"
    wb.save(path)
    wb2 = reopen_check(path)
    assert wb2["Loan"]["B1"].value == 24000
    save_manifest(tid, [{"filename": "loan.xlsx", "sandbox_path": "/home/user/Documents/loan.xlsx", "type": "xlsx"}])
    print(f"[{tid}] ok")


# ----- 2. calc_sumifs_multi_store -----
def gen_sumifs():
    tid = "calc_sumifs_multi_store"
    env = ensure(tid)
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    headers = ["Date", "Store", "Category", "Amount"]
    for i, h in enumerate(headers):
        c = ws.cell(row=1, column=1 + i, value=h); c.font = BOLD
    # 40 rows deterministic
    stores = ["North", "South", "West"]
    cats = ["Food", "Clothing", "Electronics", "Home"]
    rows = []
    # fixed data
    amounts = [120, 85, 300, 55, 210, 175, 40, 90, 260, 130,
               95, 310, 65, 155, 220, 80, 115, 70, 190, 45,
               330, 60, 140, 100, 75, 250, 165, 110, 205, 150,
               50, 35, 280, 125, 180, 95, 240, 290, 170, 200]
    assert len(amounts) == 40
    idx = 0
    for i in range(40):
        store = stores[i % 3]
        cat = cats[(i // 3) % 4]
        amt = amounts[i]
        ws.append([f"2025-03-{(i % 28) + 1:02d}", store, cat, amt])
        rows.append((store, cat, amt))
    path = env / "transactions.xlsx"
    wb.save(path)
    # compute expected totals per store+category and per store
    totals = {s: {c: 0 for c in cats} for s in stores}
    for s, c, a in rows:
        totals[s][c] += a
    print(f"[{tid}] totals", totals)
    for s in stores:
        row_total = sum(totals[s].values())
        print(f"  {s} total = {row_total}")
    cat_totals = {c: sum(totals[s][c] for s in stores) for c in cats}
    print(f"  category totals = {cat_totals}")
    print(f"  grand total = {sum(amounts)}")
    wb2 = reopen_check(path)
    assert wb2["Transactions"]["A1"].value == "Date"
    save_manifest(tid, [{"filename": "transactions.xlsx", "sandbox_path": "/home/user/Documents/transactions.xlsx", "type": "xlsx"}])


# ----- 3. calc_month_over_month_growth -----
def gen_mom():
    tid = "calc_month_over_month_growth"
    env = ensure(tid)
    wb = Workbook()
    ws = wb.active
    ws.title = "Revenue"
    ws["A1"] = "Month"; ws["B1"] = "Revenue"
    ws["A1"].font = BOLD; ws["B1"].font = BOLD
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    revs = [100000, 110000, 105000, 120000, 125000, 130000,
            128000, 140000, 135000, 150000, 160000, 155000]
    for i, (m, r) in enumerate(zip(months, revs)):
        ws.cell(row=2 + i, column=1, value=m)
        ws.cell(row=2 + i, column=2, value=r)
    path = env / "revenue.xlsx"
    wb.save(path)
    # expected
    running = []
    s = 0
    for r in revs:
        s += r; running.append(s)
    print(f"[{tid}] running totals", running)
    mom = ["N/A"]
    for i in range(1, 12):
        pct = (revs[i] - revs[i - 1]) / revs[i - 1] * 100
        mom.append(round(pct, 6))
    print(f"  mom growth", mom)
    status = ["N/A"] + ["Down" if isinstance(x, float) and x < 0 else "Up" for x in mom[1:]]
    print(f"  status", status)
    wb2 = reopen_check(path)
    assert wb2["Revenue"]["B2"].value == 100000
    save_manifest(tid, [{"filename": "revenue.xlsx", "sandbox_path": "/home/user/Documents/revenue.xlsx", "type": "xlsx"}])


# ----- 4. calc_index_match_lookup -----
def gen_index_match():
    tid = "calc_index_match_lookup"
    env = ensure(tid)
    wb = Workbook()
    cat = wb.active
    cat.title = "Catalog"
    cat.append(["SKU", "Product", "Unit Price"])
    for c in ("A1", "B1", "C1"):
        cat[c].font = BOLD
    products = [
        ("P001", "Widget A", 12.50),
        ("P002", "Widget B", 15.00),
        ("P003", "Gadget X", 22.75),
        ("P004", "Gadget Y", 30.00),
        ("P005", "Thingamajig", 8.25),
        ("P006", "Doohickey", 18.40),
        ("P007", "Gizmo", 27.10),
        ("P008", "Contraption", 45.00),
        ("P009", "Apparatus", 33.60),
        ("P010", "Device", 19.99),
    ]
    for sku, name, price in products:
        cat.append([sku, name, price])
    q = wb.create_sheet("Quotes")
    q.append(["SKU", "Qty"])
    q["A1"].font = BOLD; q["B1"].font = BOLD
    quotes = [
        ("P003", 5), ("P001", 10), ("P008", 2), ("P005", 20), ("P007", 3),
        ("P010", 8), ("P002", 4), ("P009", 6), ("P006", 12), ("P004", 7),
    ]
    for sku, qty in quotes:
        q.append([sku, qty])
    path = env / "quotes.xlsx"
    wb.save(path)
    # compute expected grand total
    price_map = {sku: price for sku, _, price in products}
    name_map = {sku: name for sku, name, _ in products}
    total = 0.0
    for sku, qty in quotes:
        total += price_map[sku] * qty
    print(f"[{tid}] grand total = {total}")
    print(f"  line 2: product={name_map['P003']}, price={price_map['P003']}, line total={price_map['P003']*5}")
    wb2 = reopen_check(path)
    assert wb2["Catalog"]["B4"].value == "Gadget X"
    save_manifest(tid, [{"filename": "quotes.xlsx", "sandbox_path": "/home/user/Documents/quotes.xlsx", "type": "xlsx"}])


# ----- 5. calc_payroll_tax_brackets -----
def gen_payroll():
    tid = "calc_payroll_tax_brackets"
    env = ensure(tid)
    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll"
    ws.append(["Employee", "Hourly Rate", "Hours Worked"])
    for c in ("A1", "B1", "C1"):
        ws[c].font = BOLD
    data = [
        ("Alice", 20, 45),
        ("Bob", 15, 40),
        ("Carol", 25, 50),
        ("Dan", 18, 38),
        ("Eve", 22, 42),
        ("Frank", 30, 55),
        ("Grace", 12, 35),
        ("Henry", 28, 40),
        ("Irene", 16, 48),
        ("Jack", 35, 30),
        ("Kate", 24, 44),
        ("Leo", 19, 46),
    ]
    for r in data:
        ws.append(list(r))
    path = env / "payroll.xlsx"
    wb.save(path)
    # compute expected
    def tax(gross):
        if gross < 500: return 0.0
        if gross <= 1500: return (gross - 500) * 0.10
        return 1000 * 0.10 + (gross - 1500) * 0.20
    sums = [0.0] * 5  # regular, ot, gross, tax, net
    for name, rate, hours in data:
        reg = min(hours, 40) * rate
        ot = max(0, hours - 40) * rate * 1.5
        gross = reg + ot
        t = tax(gross)
        net = gross - t
        print(f"[{tid}] {name}: reg={reg} ot={ot} gross={gross} tax={t} net={net}")
        sums[0] += reg; sums[1] += ot; sums[2] += gross; sums[3] += t; sums[4] += net
    print(f"  totals reg={sums[0]} ot={sums[1]} gross={sums[2]} tax={sums[3]} net={sums[4]}")
    wb2 = reopen_check(path)
    assert wb2["Payroll"]["C2"].value == 45
    save_manifest(tid, [{"filename": "payroll.xlsx", "sandbox_path": "/home/user/Documents/payroll.xlsx", "type": "xlsx"}])


# ----- 6. calc_inventory_reorder_flags -----
def gen_reorder():
    tid = "calc_inventory_reorder_flags"
    env = ensure(tid)
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock"
    ws.append(["SKU", "Item", "On Hand", "Daily Usage", "Reorder Point", "Max Stock"])
    for col in "ABCDEF":
        ws[f"{col}1"].font = BOLD
    data = [
        ("S001", "Bolts", 30, 5, 50, 200),      # Low
        ("S002", "Nuts", 10, 4, 40, 180),       # Critical (<20)
        ("S003", "Washers", 150, 10, 80, 300),  # OK
        ("S004", "Screws", 5, 2, 30, 150),      # Critical (<15)
        ("S005", "Nails", 60, 6, 70, 250),      # Low
        ("S006", "Hinges", 80, 4, 40, 200),     # OK
        ("S007", "Brackets", 12, 3, 30, 120),   # Critical (<15)
        ("S008", "Clamps", 25, 2, 20, 100),     # OK
        ("S009", "Pins", 18, 5, 40, 160),       # Critical (<20)
        ("S010", "Rods", 45, 4, 60, 220),       # Low
        ("S011", "Plates", 200, 8, 90, 400),    # OK
        ("S012", "Springs", 3, 1, 15, 80),      # Critical (<7.5)
        ("S013", "Rings", 55, 3, 60, 180),      # Low
        ("S014", "Caps", 110, 5, 50, 250),      # OK
        ("S015", "Valves", 7, 2, 20, 90),       # Critical (<10)
    ]
    for r in data:
        ws.append(list(r))
    path = env / "stock.xlsx"
    wb.save(path)
    counts = {"Critical": 0, "Low": 0, "OK": 0}
    for sku, item, oh, du, rp, mx in data:
        if oh < rp / 2:
            status = "Critical"
        elif oh < rp:
            status = "Low"
        else:
            status = "OK"
        reorder = (mx - oh) if status in ("Critical", "Low") else 0
        dos = oh / du
        counts[status] += 1
        print(f"[{tid}] {sku} oh={oh} dos={dos} status={status} reorder={reorder}")
    print(f"  counts {counts}")
    wb2 = reopen_check(path)
    assert wb2["Stock"]["B2"].value == "Bolts"
    save_manifest(tid, [{"filename": "stock.xlsx", "sandbox_path": "/home/user/Documents/stock.xlsx", "type": "xlsx"}])


# ----- 7. calc_date_analysis_workdays -----
def gen_dates():
    tid = "calc_date_analysis_workdays"
    env = ensure(tid)
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    ws.append(["Order ID", "Order Date", "Amount"])
    for c in "ABC":
        ws[f"{c}1"].font = BOLD
    from datetime import date
    data = [
        (1001, date(2025, 1, 3), 120),
        (1002, date(2025, 1, 5), 80),   # Sunday
        (1003, date(2025, 1, 10), 200),
        (1004, date(2025, 1, 15), 150),
        (1005, date(2025, 1, 20), 95),
        (1006, date(2025, 1, 25), 175), # Saturday
        (1007, date(2025, 1, 28), 60),
        (1008, date(2025, 2, 3), 220),
        (1009, date(2025, 2, 8), 110),  # Saturday
        (1010, date(2025, 2, 12), 140),
        (1011, date(2025, 2, 17), 85),
        (1012, date(2025, 2, 22), 195), # Saturday
        (1013, date(2025, 2, 27), 70),
        (1014, date(2025, 3, 4), 160),
        (1015, date(2025, 3, 9), 125),  # Sunday
        (1016, date(2025, 3, 14), 90),
        (1017, date(2025, 3, 18), 240),
        (1018, date(2025, 3, 22), 55),  # Saturday
        (1019, date(2025, 3, 25), 185),
        (1020, date(2025, 3, 30), 100), # Sunday
    ]
    for oid, d, amt in data:
        ws.append([oid, d, amt])
    path = env / "orders.xlsx"
    wb.save(path)
    months = {1: 0, 2: 0, 3: 0}
    for _, d, amt in data:
        months[d.month] += amt
    print(f"[{tid}] monthly totals", months)
    wb2 = reopen_check(path)
    assert wb2["Orders"]["A2"].value == 1001
    save_manifest(tid, [{"filename": "orders.xlsx", "sandbox_path": "/home/user/Documents/orders.xlsx", "type": "xlsx"}])


# ----- 8. calc_text_parse_contacts -----
def gen_contacts():
    tid = "calc_text_parse_contacts"
    env = ensure(tid)
    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"
    ws["A1"] = "Full Entry"; ws["A1"].font = BOLD
    data = [
        ("Smith", "Jane", "jane.smith@acme.io"),
        ("Brown", "Alex", "alex.brown@example.com"),
        ("Johnson", "Emily", "emily.j@mail.net"),
        ("Lee", "Marcus", "marcus.lee@globex.co"),
        ("Patel", "Priya", "priya.p@initech.biz"),
        ("Nguyen", "Kim", "kim.n@hooli.io"),
        ("Garcia", "Diego", "diego.g@umbrella.com"),
        ("Wilson", "Rachel", "rachel.w@stark.co"),
        ("Clark", "Tom", "tom.c@wayne.net"),
        ("Adams", "Nora", "nora.a@cyberdyne.com"),
        ("Turner", "Ian", "ian.t@tyrell.io"),
        ("Mitchell", "Laura", "laura.m@weyland.biz"),
        ("Scott", "Ben", "ben.s@soylent.co"),
        ("Hall", "Zoe", "zoe.h@oscorp.com"),
        ("Young", "Chris", "chris.y@lexcorp.net"),
    ]
    for ln, fn, email in data:
        ws.append([f"{ln}, {fn} <{email}>"])
    path = env / "contacts.xlsx"
    wb.save(path)
    print(f"[{tid}] first = {data[0]}")
    wb2 = reopen_check(path)
    assert "Smith" in wb2["Contacts"]["A2"].value
    save_manifest(tid, [{"filename": "contacts.xlsx", "sandbox_path": "/home/user/Documents/contacts.xlsx", "type": "xlsx"}])


# ----- 9. calc_3d_quarterly_consolidation -----
def gen_3d():
    tid = "calc_3d_quarterly_consolidation"
    env = ensure(tid)
    wb = Workbook()
    wb.remove(wb.active)
    products = ["Alpha", "Beta", "Gamma", "Delta", "Epsilon", "Zeta", "Eta", "Theta"]
    q_sales = {
        "Q1": [12000, 8000, 15000, 6000, 20000, 9000, 14000, 11000],
        "Q2": [13500, 8500, 16000, 6500, 21000, 9500, 15000, 11500],
        "Q3": [14000, 9000, 17000, 7000, 22500, 10000, 16000, 12500],
        "Q4": [15500, 9800, 18000, 7500, 24000, 10500, 17000, 13500],
    }
    prior_year = [50000, 32000, 58000, 25000, 80000, 36000, 55000, 42000]
    for qname in ["Q1", "Q2", "Q3", "Q4"]:
        s = wb.create_sheet(qname)
        s.append(["Product", "Sales"])
        s["A1"].font = BOLD; s["B1"].font = BOLD
        for p, v in zip(products, q_sales[qname]):
            s.append([p, v])
    py = wb.create_sheet("PriorYear")
    py.append(["Product", "Prior"])
    py["A1"].font = BOLD; py["B1"].font = BOLD
    for p, v in zip(products, prior_year):
        py.append([p, v])
    path = env / "quarters.xlsx"
    wb.save(path)
    annual = [sum(q_sales[q][i] for q in ["Q1", "Q2", "Q3", "Q4"]) for i in range(8)]
    print(f"[{tid}] annual = {annual}")
    print(f"  total = {sum(annual)}")
    for i, p in enumerate(products):
        yoy = (annual[i] - prior_year[i]) / prior_year[i] * 100
        print(f"  {p}: annual={annual[i]} prior={prior_year[i]} yoy%={yoy:.2f}")
    wb2 = reopen_check(path)
    assert wb2["Q1"]["B2"].value == 12000
    save_manifest(tid, [{"filename": "quarters.xlsx", "sandbox_path": "/home/user/Documents/quarters.xlsx", "type": "xlsx"}])


# ----- 10. calc_competition_rank_percentile -----
def gen_rank():
    tid = "calc_competition_rank_percentile"
    env = ensure(tid)
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append(["Name", "Country", "Score"])
    for c in "ABC":
        ws[f"{c}1"].font = BOLD
    data = [
        ("Ana", "BRA", 87), ("Bill", "USA", 92), ("Chen", "CHN", 78),
        ("Dora", "GER", 95), ("Elin", "SWE", 81), ("Femi", "NGA", 74),
        ("Gia", "ITA", 89), ("Haru", "JPN", 98), ("Ivan", "RUS", 76),
        ("Jade", "CAN", 84), ("Kira", "KOR", 90), ("Luka", "CRO", 72),
        ("Maya", "IND", 85), ("Nora", "NOR", 80), ("Omar", "EGY", 77),
        ("Pia", "ESP", 91), ("Quin", "AUS", 83), ("Ravi", "IND", 79),
        ("Sara", "FRA", 86), ("Toni", "MEX", 73),
    ]
    for r in data:
        ws.append(list(r))
    path = env / "competition.xlsx"
    wb.save(path)
    # compute ranks (descending by score)
    sorted_by = sorted(enumerate(data), key=lambda x: -x[1][2])
    rank_map = {}
    for pos, (orig, _) in enumerate(sorted_by):
        rank_map[orig] = pos + 1
    for i, (n, c, s) in enumerate(data):
        r = rank_map[i]
        medal = ""
        if r == 1: medal = "Gold"
        elif r == 2: medal = "Silver"
        elif r == 3: medal = "Bronze"
        elif r <= 5: medal = "Finalist"
        print(f"[{tid}] {n} score={s} rank={r} medal={medal}")
    # top country
    top_i = [i for i, r in rank_map.items() if r == 1][0]
    print(f"  top country = {data[top_i][1]} ({data[top_i][0]})")
    wb2 = reopen_check(path)
    assert wb2["Results"]["A2"].value == "Ana"
    save_manifest(tid, [{"filename": "competition.xlsx", "sandbox_path": "/home/user/Documents/competition.xlsx", "type": "xlsx"}])


def main():
    gen_loan()
    gen_sumifs()
    gen_mom()
    gen_index_match()
    gen_payroll()
    gen_reorder()
    gen_dates()
    gen_contacts()
    gen_3d()
    gen_rank()
    print("all env generated")


if __name__ == "__main__":
    main()
