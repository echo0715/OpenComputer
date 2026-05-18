"""Generate env files for the 4 new LibreOffice Calc v2 (round 2) tasks:
- calc_pref_default_sheets_and_save_format/env/blank.xlsx (placeholder)
- calc_conditional_formatting_sales_heatmap/env/regional_sales.xlsx
- calc_named_ranges_and_data_validation/env/orders.xlsx
- calc_autofilter_and_csv_export/env/employees.xlsx
"""

import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font

ROOT = os.path.dirname(os.path.abspath(__file__))
BOLD = Font(bold=True)


def _bold_header(ws, headers):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = BOLD


def make_blank():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "placeholder"
    out = os.path.join(ROOT, "calc_pref_default_sheets_and_save_format/env/blank.xlsx")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    return out


def make_regional_sales():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    _bold_header(ws, ["Region", "Q1", "Q2", "Q3", "Q4", "Total"])
    regions = [
        ("North",    25000, 28000, 30000, 27000),   # 110000 (>100k)
        ("South",    18000, 19000, 20000, 21000),   # 78000
        ("East",     15000, 16000, 17000, 18000),   # 66000
        ("West",     32000, 30000, 28000, 35000),   # 125000 (>100k)
        ("Central",  22000, 24000, 23000, 25000),   # 94000
        ("NE",       12000, 13000, 14000, 15000),   # 54000
        ("NW",       29000, 27000, 30000, 28000),   # 114000 (>100k)
        ("SE",        9000, 10000, 11000, 12000),   # 42000
        ("SW",       20000, 22000, 24000, 26000),   # 92000
        ("Pacific",   8000,  9500,  9000, 10500),   # 37000
        ("Mountain", 11000, 12000, 13000, 12000),   # 48000
        ("Atlantic", 30000, 32000, 33000, 31000),   # 126000 (>100k)
    ]
    for i, row in enumerate(regions, start=2):
        ws.cell(row=i, column=1, value=row[0])
        for j, v in enumerate(row[1:], start=2):
            ws.cell(row=i, column=j, value=v)
        ws.cell(row=i, column=6, value=f"=SUM(B{i}:E{i})")
    out = os.path.join(ROOT, "calc_conditional_formatting_sales_heatmap/env/regional_sales.xlsx")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    return out


def make_orders():
    wb = Workbook()
    orders = wb.active
    orders.title = "Orders"
    _bold_header(orders, ["Order ID", "Customer", "Item", "Qty", "Price", "Total w/ Tax"])
    items = [
        "Widget A", "Widget B", "Gadget X", "Gadget Y", "Sprocket",
        "Bolt Pack", "Screw Pack", "Nail Pack", "Hinge", "Bracket",
    ]
    for i in range(20):
        row = i + 2
        orders.cell(row=row, column=1, value=f"ORD-{i+1:03d}")
        # column B left blank by design
        orders.cell(row=row, column=3, value=items[i % len(items)])
        orders.cell(row=row, column=4, value=(i % 5) + 1)
        orders.cell(row=row, column=5, value=float(10 + i * 1.5))

    customers = wb.create_sheet("Customers")
    _bold_header(customers, ["Customer Name"])
    names = [
        "Acme Corp", "BrightWorks", "Cedar Industries", "Delta Supply",
        "Echo Ltd", "Fjord Materials", "Globex", "Helix Labs", "Ikonix",
    ]
    for i, nm in enumerate(names, start=2):
        customers.cell(row=i, column=1, value=nm)

    settings = wb.create_sheet("Settings")
    c = settings.cell(row=1, column=1, value="Tax Rate")
    c.font = BOLD
    settings.cell(row=1, column=2, value=0.08)

    out = os.path.join(ROOT, "calc_named_ranges_and_data_validation/env/orders.xlsx")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    return out


def make_employees():
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    _bold_header(ws, ["EmployeeID", "Name", "Department", "Salary", "HireDate"])
    data = [
        (1001, "Alice Abbott",   "Engineering",  95000, date(2019,  3, 14)),  # >80k
        (1002, "Bob Brown",      "Sales",        62000, date(2020,  7,  1)),
        (1003, "Carol Chen",     "Engineering",  78000, date(2018, 11, 23)),
        (1004, "David Diaz",     "Marketing",    55000, date(2021,  4, 10)),
        (1005, "Eva Edwards",    "Engineering", 110000, date(2015,  1,  8)),  # >80k
        (1006, "Frank Fisher",   "HR",           48000, date(2022,  6, 15)),
        (1007, "Grace Gomez",    "Finance",      72000, date(2019,  9,  2)),
        (1008, "Henry Huang",    "Engineering",  88000, date(2017, 10, 30)),  # >80k
        (1009, "Iris Ivanov",    "Sales",        51000, date(2023,  2, 19)),
        (1010, "Jack Jones",     "Finance",      66000, date(2020,  5,  5)),
        (1011, "Kara Klein",     "Marketing",    58000, date(2021,  8, 22)),
        (1012, "Leo Lin",        "Engineering", 102000, date(2016, 12,  1)),  # >80k
        (1013, "Mia Martinez",   "Sales",        47000, date(2022,  9, 16)),
        (1014, "Noah Nguyen",    "HR",           53000, date(2020,  3, 27)),
        (1015, "Olivia Owens",   "Finance",      75000, date(2018,  6,  4)),
        (1016, "Peter Park",     "Engineering",  90000, date(2017,  2, 11)),  # >80k
        (1017, "Quinn Quintana", "Marketing",    49000, date(2023,  1, 25)),
        (1018, "Ryan Reed",      "Sales",        61000, date(2019, 12, 12)),
        (1019, "Sara Singh",     "Finance",      79000, date(2020, 10,  3)),
        (1020, "Tom Thompson",   "HR",           52000, date(2022, 11, 29)),
    ]
    for i, row in enumerate(data, start=2):
        for j, v in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=v)
    out = os.path.join(ROOT, "calc_autofilter_and_csv_export/env/employees.xlsx")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    return out


if __name__ == "__main__":
    for fn in (make_blank, make_regional_sales, make_orders, make_employees):
        path = fn()
        size = os.path.getsize(path)
        print(f"wrote {path} ({size} bytes)")
